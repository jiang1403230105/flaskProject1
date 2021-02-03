#解析excel表
from openpyxl import Workbook,load_workbook
import datetime

filename='D:\jyj\国家统计局数据\快速查询.xlsx'

# wb = openpyxl.load_workbook(filename)
#
# print(wb.sheetnames)

# wb=Workbook()
# ws=wb.active
# ws['A1']=datetime.datetime.now().strftime("%Y-%m-%d")
# ws1=wb.create_sheet("jyjsheet")
# ws2=wb.create_sheet("sheet",0)
# print(wb.sheetnames)
# d=ws.cell(row=4,column=2,value=100)
# print(d)
# wb.save("jyj.xlsx")
wb2=load_workbook(filename)
active_file=wb2.active
#读取A5单元格的内容
a1=active_file['A5'].value
print(a1)
#读取最大行
row_max=active_file.max_row
print(row_max)
#读取最小行
row_min=active_file.min_row
print(row_min)
#读取最大列
column_max=active_file.max_column
print(column_max)
#读取最小列
column_min=active_file.min_column
print(column_min)
for i in range(row_min,row_max+1):
    for j in range(column_min,column_max+1):
        print(active_file.cell(row=i,column=j).value,end="\t")
    print("")